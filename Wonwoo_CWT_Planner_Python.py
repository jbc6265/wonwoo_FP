from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
import os
import subprocess
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook


APP_TITLE = "원우ENG CWT 생산계획 수립 프로그램"
APP_VERSION = "20260722_2135"
EXCEL_FILETYPES = [("Excel Workbook", "*.xlsx *.xlsm"), ("All files", "*.*")]

MONTHLY_REQUIRED = {
    "production_no": "생산번호",
    "sales_model": "영업모델",
    "machine_no": "차대호기",
    "seq": "순번",
    "serial": "연번",
    "line": "라인",
    "start_date": "착수일",
    "cwt": "CWT",
    "design_model": "설계모델",
    "remark": "비고",
}

MATERIAL_REQUIRED = {
    "logistics_no": "물류번호",
    "material_no": "자재번호",
    "material_name": "품명",
    "required_qty": "소요량",
    "order_qty": "발주량",
    "po_no": "발주번호",
    "due_date": "납기일자",
}

PLAN_COLUMNS = [
    "착수일",
    "라인",
    "생산번호(물류번호)",
    "자재번호",
    "품명",
    "모델명",
    "차대호기",
    "순번",
    "연번",
    "CWT",
    "RADAR",
    "설계모델",
    "소요량",
    "발주량",
    "발주번호",
    "납기일자",
    "비고",
]

EXCEPTION_COLUMNS = ["예외유형", "라인", "생산번호/물류번호", "영업모델", "자재번호", "사유"]


@dataclass
class SourceRows:
    file_name: str
    sheet_name: str
    headers: list[str]
    rows: list[dict]
    header_row: int


@dataclass
class PlanResult:
    plans: list[dict]
    exceptions: list[dict]
    summary: list[list]
    output_path: Path
    period_label: str
    line1_count: int
    line2_count: int


def normalize_header(value) -> str:
    return str(value or "").replace(" ", "").strip()


def normalize_key(value) -> str:
    return str(value or "").replace(" ", "").strip()


def normalize_search_text(value) -> str:
    return str(value or "").replace(" ", "").upper()


def value_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def date_text(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
        except Exception:
            pass

    raw = value_text(value)
    if not raw:
        return ""

    for sep in ("-", ".", "/"):
        parts = raw.split(sep)
        if len(parts) >= 3 and len(parts[0]) == 4:
            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2][:2].zfill(2)}"

    if "년" in raw and "월" in raw and "일" in raw:
        cleaned = raw.replace("년", "-").replace("월", "-").replace("일", "")
        parts = [part.strip() for part in cleaned.split("-")]
        if len(parts) >= 3:
            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"

    return raw


def safe_int(value) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return 999999


def find_col(headers: list[str], names: list[str], *, last: bool = False) -> int | None:
    targets = {normalize_header(name).upper() for name in names}
    matches = [index for index, header in enumerate(headers) if normalize_header(header).upper() in targets]
    if not matches:
        return None
    return matches[-1] if last else matches[0]


def score_header_row(values: list, required_headers: list[str]) -> int:
    cells = {normalize_header(value).upper() for value in values}
    return sum(1 for header in required_headers if normalize_header(header).upper() in cells)


def detect_header_row(ws, preferred_row: int, required_headers: list[str]) -> int:
    best_row = preferred_row
    best_score = score_header_row([cell.value for cell in ws[preferred_row]], required_headers)
    scan_limit = min(ws.max_row, 12)

    for row_no in range(1, scan_limit + 1):
        score = score_header_row([cell.value for cell in ws[row_no]], required_headers)
        if score > best_score:
            best_row = row_no
            best_score = score

    return best_row if best_score >= min(2, len(required_headers)) else preferred_row


def read_excel_rows(path: Path, preferred_header_row: int, required_headers: list[str]) -> SourceRows:
    validate_excel_path(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    ws = workbook.active
    sheet_name = ws.title
    header_row = detect_header_row(ws, preferred_header_row, required_headers)
    headers = [normalize_header(cell.value) for cell in ws[header_row]]
    rows = []

    for row_no, row_values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        values = list(row_values)
        if not any(value_text(value) for value in values):
            continue
        rows.append({"row_number": row_no, "values": values})

    workbook.close()
    return SourceRows(path.name, sheet_name, headers, rows, header_row)


def validate_excel_path(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(".xlsx 또는 .xlsm 파일만 지원합니다. .xls 파일은 xlsx로 저장 후 다시 선택하세요.")


def value_at(row: dict, index: int | None) -> str:
    if index is None:
        return ""
    values = row["values"]
    if index >= len(values):
        return ""
    return value_text(values[index])


def date_at(row: dict, index: int | None) -> str:
    if index is None:
        return ""
    values = row["values"]
    if index >= len(values):
        return ""
    return date_text(values[index])


def build_monthly_map(headers: list[str]) -> dict[str, int | None]:
    return {
        "production_no": find_col(headers, ["생산번호", "물류번호"]),
        "sales_model": find_col(headers, ["영업모델"]),
        "machine_no": find_col(headers, ["차대호기"]),
        "seq": find_col(headers, ["순번"], last=True),
        "serial": find_col(headers, ["연번"]),
        "line": find_col(headers, ["라인"]),
        "start_date": find_col(headers, ["착수일", "착수일자"]),
        "radar": find_col(headers, ["RADAR", "Radar"]),
        "cwt": find_col(headers, ["CWT"]),
        "design_model": find_col(headers, ["설계모델"]),
        "remark": find_col(headers, ["비고"]),
    }


def build_material_map(headers: list[str]) -> dict[str, int | None]:
    mapping = {
        "logistics_no": find_col(headers, ["물류번호", "생산번호"]),
        "material_no": find_col(headers, ["자재번호"]),
        "material_name": find_col(headers, ["품명"]),
        "required_qty": find_col(headers, ["소요량"]),
        "order_qty": find_col(headers, ["발주량"]),
        "po_no": find_col(headers, ["발주번호"]),
        "due_date": find_col(headers, ["납기일자", "납품예정일"]),
    }
    if len(headers) >= 24:
        fallback = {
            "logistics_no": 3,
            "material_no": 4,
            "material_name": 5,
            "required_qty": 8,
            "order_qty": 9,
            "po_no": 20,
            "due_date": 23,
        }
        for key, index in fallback.items():
            if mapping[key] is None:
                mapping[key] = index
    return mapping


def validate_columns(label: str, mapping: dict[str, int | None], required: dict[str, str], exceptions: list[dict]) -> None:
    missing = [name for key, name in required.items() if mapping.get(key) is None]
    if missing:
        exceptions.append(to_exception("필수 컬럼 누락", label, "", "", "", f"{label}: {', '.join(missing)} 컬럼을 찾을 수 없습니다."))


def make_monthly(row: dict, mapping: dict[str, int | None], line_type: str) -> dict:
    return {
        "production_no": normalize_key(value_at(row, mapping["production_no"])),
        "sales_model": value_at(row, mapping["sales_model"]),
        "machine_no": value_at(row, mapping["machine_no"]),
        "seq": value_at(row, mapping["seq"]),
        "serial": value_at(row, mapping["serial"]),
        "line": value_at(row, mapping["line"]),
        "line_type": line_type,
        "start_date": date_at(row, mapping["start_date"]),
        "radar": value_at(row, mapping["radar"]),
        "cwt": value_at(row, mapping["cwt"]),
        "design_model": value_at(row, mapping["design_model"]),
        "remark": value_at(row, mapping["remark"]),
    }


def parse_monthly_rows(source: SourceRows, mapping: dict[str, int | None], line_type: str, exceptions: list[dict]) -> list[dict]:
    parsed = []
    for row in source.rows:
        item = make_monthly(row, mapping, line_type)
        if not item["production_no"]:
            exceptions.append(to_exception("필수 값 누락", line_type, "", item["sales_model"], "", f"{source.file_name} {row['row_number']}행 생산번호가 비어 있습니다."))
            continue
        parsed.append(item)
    return parsed


def parse_line2_rows(source: SourceRows, mapping: dict[str, int | None], exceptions: list[dict]) -> tuple[list[dict], list[dict]]:
    hyundai = []
    dx = []
    for row in source.rows:
        item = make_monthly(row, mapping, "통합2라인")
        if not item["production_no"]:
            continue
        sales_model = str(item["sales_model"]).strip()
        if sales_model.startswith(("HX", "R")):
            hyundai.append(item)
        elif sales_model.startswith("DX"):
            dx.append(item)
            exceptions.append(to_exception("DX 제품 제외", "통합2라인", item["production_no"], sales_model, "", "통합2라인 디벨론 제품은 생산계획 I/F 대상에서 제외합니다."))
        else:
            exceptions.append(to_exception("모델 필터 제외", "통합2라인", item["production_no"], sales_model, "", "통합2라인 현대 제품 기준 HX~/R~에 해당하지 않습니다."))
    return hyundai, dx


def parse_material_rows(source: SourceRows, mapping: dict[str, int | None]) -> list[dict]:
    rows = []
    for row in source.rows:
        logistics_no = normalize_key(value_at(row, mapping["logistics_no"]))
        material_name = value_at(row, mapping["material_name"])
        if not logistics_no:
            continue
        rows.append({
            "logistics_no": logistics_no,
            "material_no": value_at(row, mapping["material_no"]),
            "material_name": material_name,
            "required_qty": value_at(row, mapping["required_qty"]),
            "order_qty": value_at(row, mapping["order_qty"]),
            "po_no": value_at(row, mapping["po_no"]),
            "due_date": date_at(row, mapping["due_date"]),
            "line_type": line_type_from_logistics(logistics_no),
        })
    return rows


def line_type_from_logistics(logistics_no: str) -> str:
    if logistics_no.startswith("KPA10"):
        return "통합1라인"
    if logistics_no.startswith("KPA20"):
        return "통합2라인"
    return "라인 미분류"


def is_counterweight_name(name: str) -> bool:
    return "카운터웨이트" in normalize_search_text(name)


def to_plan(monthly: dict, material: dict) -> dict:
    return {
        "착수일": monthly["start_date"],
        "라인": monthly["line_type"],
        "생산번호(물류번호)": monthly["production_no"],
        "자재번호": material["material_no"],
        "품명": material["material_name"],
        "모델명": monthly["sales_model"],
        "차대호기": monthly["machine_no"],
        "순번": monthly["seq"],
        "연번": monthly["serial"],
        "CWT": monthly["cwt"],
        "RADAR": monthly["radar"],
        "설계모델": monthly["design_model"],
        "소요량": material["required_qty"],
        "발주량": material["order_qty"],
        "발주번호": material["po_no"],
        "납기일자": material["due_date"],
        "비고": monthly["remark"],
    }


def to_exception(kind: str, line: str, key: str, model: str, material_no: str, reason: str) -> dict:
    return {
        "예외유형": kind,
        "라인": line,
        "생산번호/물류번호": key,
        "영업모델": model,
        "자재번호": material_no,
        "사유": reason,
    }


def compare_plan(row: dict) -> tuple:
    return (
        str(row["착수일"]),
        str(row["라인"]),
        safe_int(row["순번"]),
        str(row["연번"]),
        str(row["생산번호(물류번호)"]),
    )


def create_output_filename(plans: list[dict]) -> tuple[str, str]:
    dates = sorted(row["착수일"] for row in plans if row["착수일"])
    if not dates:
        return "원우ENG_CWT_생산계획_START-END.xlsx", "착수일 없음"
    start = dates[0].replace("-", "")
    end = dates[-1].replace("-", "")
    return f"원우ENG_CWT_생산계획_{start}-{end}.xlsx", f"{dates[0]} ~ {dates[-1]}"


def process_files(line1_path: Path | None, line2_path: Path | None, material_path: Path, output_dir: Path) -> PlanResult:
    if not line1_path and not line2_path:
        raise ValueError("월확정 통합1라인 또는 통합2라인 파일 중 하나 이상을 선택하세요.")

    exceptions: list[dict] = []
    line1_source = read_excel_rows(line1_path, 2, ["생산번호", "영업모델", "착수일"]) if line1_path else None
    line2_source = read_excel_rows(line2_path, 2, ["생산번호", "영업모델", "착수일"]) if line2_path else None
    material_source = read_excel_rows(material_path, 1, ["물류번호", "자재번호", "품명"])

    line1_map = build_monthly_map(line1_source.headers) if line1_source else None
    line2_map = build_monthly_map(line2_source.headers) if line2_source else None
    material_map = build_material_map(material_source.headers)

    if line1_map:
        validate_columns("월확정 통합1라인", line1_map, MONTHLY_REQUIRED, exceptions)
    if line2_map:
        validate_columns("월확정 통합2라인", line2_map, MONTHLY_REQUIRED, exceptions)
    validate_columns("물류번호별 자재소요현황", material_map, MATERIAL_REQUIRED, exceptions)

    line1 = parse_monthly_rows(line1_source, line1_map, "통합1라인", exceptions) if line1_source and line1_map else []
    line2, line2_dx = parse_line2_rows(line2_source, line2_map, exceptions) if line2_source and line2_map else ([], [])
    material_all = parse_material_rows(material_source, material_map)
    material_line1 = [row for row in material_all if row["logistics_no"].startswith("KPA10") and is_counterweight_name(row["material_name"])]
    material_line2 = [row for row in material_all if row["logistics_no"].startswith("KPA20") and is_counterweight_name(row["material_name"])]
    selected_materials = []
    excluded_materials = []
    if line1_source:
        selected_materials.extend(material_line1)
    else:
        excluded_materials.extend(material_line1)
    if line2_source:
        selected_materials.extend(material_line2)
    else:
        excluded_materials.extend(material_line2)

    material_index: dict[str, list[dict]] = defaultdict(list)
    all_material_index: dict[str, list[dict]] = defaultdict(list)
    for item in material_all:
        all_material_index[item["logistics_no"]].append(item)
    for item in selected_materials:
        material_index[item["logistics_no"]].append(item)

    plans = []
    monthly = [*line1, *line2]
    for item in monthly:
        matches = material_index.get(item["production_no"], [])
        if len(matches) == 1:
            plans.append(to_plan(item, matches[0]))
        elif len(matches) == 0:
            candidates = all_material_index.get(item["production_no"], [])
            if candidates:
                material_nos = ", ".join(row["material_no"] for row in candidates if row["material_no"])
                names = " / ".join(sorted({row["material_name"] for row in candidates if row["material_name"]}))
                exceptions.append(to_exception("카운터웨이트 품명 미확인", item["line_type"], item["production_no"], item["sales_model"], material_nos, f"동일 물류번호 자재는 있으나 품명에서 카운터웨이트를 확인하지 못했습니다. 품명: {names}"))
            else:
                exceptions.append(to_exception("자재소요 미매칭", item["line_type"], item["production_no"], item["sales_model"], "", "생산번호와 일치하는 카운터웨이트 자재소요가 없습니다."))
        else:
            material_nos = ", ".join(row["material_no"] for row in matches if row["material_no"])
            exceptions.append(to_exception("자재소요 중복", item["line_type"], item["production_no"], item["sales_model"], material_nos, "동일 생산번호에 카운터웨이트 자재소요가 2건 이상입니다."))

    monthly_keys = {row["production_no"] for row in monthly}
    for item in selected_materials:
        if item["logistics_no"] not in monthly_keys:
            exceptions.append(to_exception("월확정 미매칭", item["line_type"], item["logistics_no"], "", item["material_no"], "물류번호와 일치하는 월확정 생산번호가 없습니다."))
    for item in excluded_materials:
        exceptions.append(to_exception("선택 라인 제외", item["line_type"], item["logistics_no"], "", item["material_no"], f"{item['line_type']} 월확정 파일이 선택되지 않아 생산계획 생성 대상에서 제외했습니다."))

    plans.sort(key=compare_plan)
    output_name, period_label = create_output_filename(plans)
    output_path = output_dir / output_name

    summary = [
        ["구분", "건수", "비고"],
        ["월확정 통합1라인 전체", len(line1_source.rows) if line1_source else 0, line1_source.file_name if line1_source else "파일 미선택"],
        ["월확정 통합1라인 사용", len(line1), "선택 시 전체 사용"],
        ["월확정 통합2라인 전체", len(line2_source.rows) if line2_source else 0, line2_source.file_name if line2_source else "파일 미선택"],
        ["월확정 통합2라인 현대 제품", len(line2), "선택 시 영업모델 HX~/R~"],
        ["월확정 통합2라인 DX 제외", len(line2_dx), "예외목록 기록"],
        ["자재소요 전체", len(material_source.rows), material_source.file_name],
        ["자재소요 KPA10 카운터웨이트", len(material_line1), "통합1라인 매칭 대상"],
        ["자재소요 KPA20 카운터웨이트", len(material_line2), "통합2라인 매칭 대상"],
        ["선택 라인 제외 자재", len(excluded_materials), "선택되지 않은 월확정 라인의 자재"],
        ["정상 생산계획", len(plans), "생산계획 시트"],
        ["예외", len(exceptions), "예외목록 시트"],
        ["착수기간", period_label, "월확정 착수일 기준"],
        ["앱 버전", APP_VERSION, "Python + openpyxl + tkinter"],
    ]

    write_workbook(output_path, plans, exceptions, summary)
    return PlanResult(plans, exceptions, summary, output_path, period_label, len(line1), len(line2))


def write_workbook(output_path: Path, plans: list[dict], exceptions: list[dict], summary: list[list]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_sheet(workbook, "생산계획", plans, PLAN_COLUMNS)
    write_sheet(workbook, "예외목록", exceptions or [to_exception("예외 없음", "", "", "", "", "예외가 없습니다.")], EXCEPTION_COLUMNS)
    write_sheet(workbook, "검증요약", summary)
    workbook.save(output_path)


def write_sheet(workbook: Workbook, title: str, rows: list[dict] | list[list], columns: list[str] | None = None) -> None:
    ws = workbook.create_sheet(title)
    if not rows:
        if columns:
            ws.append(columns)
        return

    if isinstance(rows[0], dict):
        headers = columns or list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
    else:
        for row in rows:
            ws.append(row)

    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = min(width, 44)


def open_in_file_manager(path: Path) -> None:
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
    except Exception:
        pass


class PlannerApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1180x760")
        self.minsize(1040, 680)
        self.files: dict[str, Path | None] = {
            "line1": None,
            "line2": None,
            "material": None,
            "output_dir": Path.home() / "Desktop",
        }
        self.last_result: PlanResult | None = None
        self._build_ui()
        self._update_generate_state()

    def _build_ui(self) -> None:
        self.configure(bg="#f5f7fb")
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background="#f5f7fb")
        style.configure("Panel.TFrame", background="#ffffff", relief="solid", borderwidth=1)
        style.configure("Title.TLabel", background="#f5f7fb", foreground="#172033", font=("Malgun Gothic", 22, "bold"))
        style.configure("Hint.TLabel", background="#f5f7fb", foreground="#697386", font=("Malgun Gothic", 10))
        style.configure("PanelTitle.TLabel", background="#ffffff", foreground="#172033", font=("Malgun Gothic", 12, "bold"))
        style.configure("Value.TLabel", background="#ffffff", foreground="#2563eb", font=("Malgun Gothic", 10, "bold"))
        style.configure("Primary.TButton", font=("Malgun Gothic", 11, "bold"))

        root = ttk.Frame(self, padding=24)
        root.pack(fill="both", expand=True)

        ttk.Label(root, text=APP_TITLE, style="Title.TLabel").pack(anchor="w")
        ttk.Label(root, text="월확정 통합1/2라인 중 하나 이상과 자재소요현황을 생산번호 기준으로 매칭해 결과 엑셀을 생성합니다.", style="Hint.TLabel").pack(anchor="w", pady=(6, 18))

        select_frame = ttk.Frame(root)
        select_frame.pack(fill="x")
        self.file_labels: dict[str, ttk.Label] = {}
        self._add_file_row(select_frame, "line1", "1. 월확정조립서열계획 통합1라인", "선택 사항 · 선택 시 KPA10 자재와 매칭", 0)
        self._add_file_row(select_frame, "line2", "2. 월확정조립서열계획 통합2라인", "선택 사항 · 선택 시 HX/R 현대 제품만 사용", 1)
        self._add_file_row(select_frame, "material", "3. 물류번호별 자재소요현황", "필수 · KPA10/KPA20 및 카운터웨이트 품명만 사용", 2)
        self._add_output_row(select_frame, 3)

        action_frame = ttk.Frame(root)
        action_frame.pack(fill="x", pady=(16, 14))
        self.generate_btn = ttk.Button(action_frame, text="생산계획 생성", style="Primary.TButton", command=self.generate)
        self.generate_btn.pack(side="left")
        ttk.Button(action_frame, text="초기화", command=self.reset).pack(side="left", padx=(8, 0))
        self.open_btn = ttk.Button(action_frame, text="결과 폴더 열기", command=self.open_output, state="disabled")
        self.open_btn.pack(side="left", padx=(8, 0))
        self.status_var = tk.StringVar(value="월확정 1개 이상과 자재소요 파일을 선택하면 생성할 수 있습니다.")
        ttk.Label(action_frame, textvariable=self.status_var, style="Hint.TLabel").pack(side="right")

        summary = ttk.Frame(root)
        summary.pack(fill="x", pady=(2, 14))
        self.metric_vars = {
            "line1": tk.StringVar(value="-"),
            "line2": tk.StringVar(value="-"),
            "plans": tk.StringVar(value="-"),
            "exceptions": tk.StringVar(value="-"),
            "period": tk.StringVar(value="-"),
        }
        self._add_metric(summary, "통합1라인", self.metric_vars["line1"], 0)
        self._add_metric(summary, "통합2라인", self.metric_vars["line2"], 1)
        self._add_metric(summary, "정상 생산계획", self.metric_vars["plans"], 2)
        self._add_metric(summary, "예외", self.metric_vars["exceptions"], 3)
        self._add_metric(summary, "착수기간", self.metric_vars["period"], 4)

        table_frame = ttk.Frame(root, style="Panel.TFrame", padding=12)
        table_frame.pack(fill="both", expand=True)
        ttk.Label(table_frame, text="생산계획 미리보기", style="PanelTitle.TLabel").pack(anchor="w", pady=(0, 8))

        columns = ["착수일", "라인", "생산번호(물류번호)", "자재번호", "모델명", "차대호기", "CWT", "RADAR"]
        self.preview = ttk.Treeview(table_frame, columns=columns, show="headings", height=12)
        for column in columns:
            self.preview.heading(column, text=column)
            self.preview.column(column, width=130, anchor="w")
        self.preview.column("착수일", width=105)
        self.preview.column("라인", width=95)
        self.preview.column("모델명", width=150)

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.preview.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.preview.xview)
        self.preview.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.preview.pack(side="left", fill="both", expand=True)
        y_scroll.pack(side="right", fill="y")
        x_scroll.pack(side="bottom", fill="x")

    def _add_file_row(self, parent: ttk.Frame, key: str, title: str, hint: str, row: int) -> None:
        panel = ttk.Frame(parent, style="Panel.TFrame", padding=14)
        panel.grid(row=row, column=0, sticky="ew", pady=5)
        parent.columnconfigure(0, weight=1)
        ttk.Label(panel, text=title, style="PanelTitle.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(panel, text=hint, background="#ffffff", foreground="#697386").grid(row=1, column=0, sticky="w", pady=(4, 0))
        label = ttk.Label(panel, text="파일을 선택하세요", style="Value.TLabel")
        label.grid(row=0, column=1, rowspan=2, sticky="ew", padx=16)
        self.file_labels[key] = label
        ttk.Button(panel, text="파일 선택", command=lambda: self.select_file(key)).grid(row=0, column=2, rowspan=2)
        panel.columnconfigure(1, weight=1)

    def _add_output_row(self, parent: ttk.Frame, row: int) -> None:
        panel = ttk.Frame(parent, style="Panel.TFrame", padding=14)
        panel.grid(row=row, column=0, sticky="ew", pady=5)
        ttk.Label(panel, text="4. 결과 저장 폴더", style="PanelTitle.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(panel, text="생성된 생산계획 엑셀을 저장할 위치", background="#ffffff", foreground="#697386").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.output_label = ttk.Label(panel, text=str(self.files["output_dir"]), style="Value.TLabel")
        self.output_label.grid(row=0, column=1, rowspan=2, sticky="ew", padx=16)
        ttk.Button(panel, text="폴더 선택", command=self.select_output_dir).grid(row=0, column=2, rowspan=2)
        panel.columnconfigure(1, weight=1)

    def _add_metric(self, parent: ttk.Frame, title: str, var: tk.StringVar, column: int) -> None:
        panel = ttk.Frame(parent, style="Panel.TFrame", padding=14)
        panel.grid(row=0, column=column, sticky="ew", padx=(0 if column == 0 else 8, 0))
        parent.columnconfigure(column, weight=1)
        ttk.Label(panel, text=title, style="PanelTitle.TLabel").pack(anchor="w")
        ttk.Label(panel, textvariable=var, style="Value.TLabel", font=("Malgun Gothic", 15, "bold")).pack(anchor="w", pady=(6, 0))

    def select_file(self, key: str) -> None:
        path = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=EXCEL_FILETYPES)
        if not path:
            return
        selected = Path(path)
        self.files[key] = selected
        self.file_labels[key].configure(text=selected.name)
        self._update_generate_state()

    def select_output_dir(self) -> None:
        path = filedialog.askdirectory(title="결과 저장 폴더 선택")
        if not path:
            return
        self.files["output_dir"] = Path(path)
        self.output_label.configure(text=path)
        self._update_generate_state()

    def _update_generate_state(self) -> None:
        has_monthly = bool(self.files["line1"] or self.files["line2"])
        ready = has_monthly and bool(self.files["material"]) and bool(self.files["output_dir"])
        if hasattr(self, "generate_btn"):
            self.generate_btn.configure(state="normal" if ready else "disabled")

    def generate(self) -> None:
        try:
            self.status_var.set("엑셀 파일을 처리하고 있습니다...")
            self.update_idletasks()
            result = process_files(
                self.files["line1"],
                self.files["line2"],
                self.files["material"],
                self.files["output_dir"],
            )
        except Exception as exc:
            self.status_var.set("처리 중 오류가 발생했습니다.")
            messagebox.showerror(APP_TITLE, str(exc))
            return

        self.last_result = result
        self.metric_vars["line1"].set(f"{result.line1_count}건")
        self.metric_vars["line2"].set(f"{result.line2_count}건")
        self.metric_vars["plans"].set(f"{len(result.plans)}건")
        self.metric_vars["exceptions"].set(f"{len(result.exceptions)}건")
        self.metric_vars["period"].set(result.period_label)
        self.status_var.set(f"완료: {result.output_path.name}")
        self.open_btn.configure(state="normal")
        self.render_preview(result.plans)
        messagebox.showinfo(APP_TITLE, f"생산계획 생성 완료\n\n정상 {len(result.plans)}건, 예외 {len(result.exceptions)}건\n\n{result.output_path}")

    def render_preview(self, plans: list[dict]) -> None:
        for item in self.preview.get_children():
            self.preview.delete(item)
        columns = self.preview["columns"]
        for row in plans[:100]:
            self.preview.insert("", "end", values=[row.get(column, "") for column in columns])

    def open_output(self) -> None:
        if self.last_result:
            open_in_file_manager(self.last_result.output_path.parent)

    def reset(self) -> None:
        for key in ("line1", "line2", "material"):
            self.files[key] = None
            self.file_labels[key].configure(text="파일을 선택하세요")
        self.last_result = None
        for var in self.metric_vars.values():
            var.set("-")
        self.status_var.set("월확정 1개 이상과 자재소요 파일을 선택하면 생성할 수 있습니다.")
        self.open_btn.configure(state="disabled")
        self.render_preview([])
        self._update_generate_state()


def main() -> None:
    app = PlannerApp()
    app.mainloop()


if __name__ == "__main__":
    main()
