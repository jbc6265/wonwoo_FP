from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from zipfile import BadZipFile
import base64
import os
import subprocess
import sys
import tempfile
import tkinter as tk
import unicodedata
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

APP_TITLE = "원우ENG 서열정보&소요자재 자동 취합 프로그램"
APP_VERSION = "20260819_seonjin_wx_included"
MONTHLY_SOURCE_KEYS = ("line1", "line2", "seonjin", "superlarge")
ALL_SOURCE_KEYS = MONTHLY_SOURCE_KEYS + ("material",)
MONTHLY_DEFAULT_COLUMNS = ("생산번호", "영업모델", "차대호기", "착수일", "CWT", "RADAR", "연번", "순번_2", "국가")
MATERIAL_DEFAULT_COLUMNS = ("착수일자", "물류번호", "자재번호", "품명", "차대모델", "차대호기", "업체명")
SOURCE_LABELS = {
    "line1": "월확정서열 통합1라인",
    "line2": "월확정서열 통합2라인",
    "seonjin": "월확정서열 선진정공",
    "superlarge": "월확정서열 초대형",
    "material": "자재소요현황",
}
EXCEPTION_COLUMNS = ["예외유형", "파일구분", "Key", "사유"]

LIGHT_COLORS = {
    "bg": "#F8FAFC",
    "surface": "#FFFFFF",
    "surface_alt": "#F1F5F9",
    "panel": "#E9EEF6",
    "border": "#DBEAFE",
    "border_strong": "#BFDBFE",
    "text": "#0F172A",
    "muted": "#64748B",
    "primary": "#1E40AF",
    "primary_hover": "#1D4ED8",
    "secondary": "#3B82F6",
    "accent": "#D97706",
    "success": "#15803D",
    "warning": "#B45309",
    "danger": "#DC2626",
    "dark": "#111827",
    "dark_panel": "#1F2937",
    "on_dark": "#FFFFFF",
    "header_sub": "#CBD5E1",
    "header_key": "#93C5FD",
}
DARK_COLORS = {
    "bg": "#0B1120",
    "surface": "#111827",
    "surface_alt": "#1E293B",
    "panel": "#243044",
    "border": "#334155",
    "border_strong": "#475569",
    "text": "#F8FAFC",
    "muted": "#CBD5E1",
    "primary": "#60A5FA",
    "primary_hover": "#3B82F6",
    "secondary": "#93C5FD",
    "accent": "#F59E0B",
    "success": "#4ADE80",
    "warning": "#FBBF24",
    "danger": "#F87171",
    "dark": "#020617",
    "dark_panel": "#0F172A",
    "on_dark": "#FFFFFF",
    "header_sub": "#D1D5DB",
    "header_key": "#BFDBFE",
}
SOURCE_TONES_LIGHT = {
    "line1": {"surface": "#F5F8FE", "header": "#EDF4FE", "chip": "#E8F0FE", "border": "#C6DAFC", "accent": "#4285F4"},
    "line2": {"surface": "#FEF6F5", "header": "#FDEEEE", "chip": "#FCE8E6", "border": "#F6C7C3", "accent": "#EA4335"},
    "seonjin": {"surface": "#FFFBF0", "header": "#FFF7E0", "chip": "#FEF3C7", "border": "#F8D98A", "accent": "#D99400"},
    "superlarge": {"surface": "#F3F9F5", "header": "#EAF6EE", "chip": "#E6F4EA", "border": "#B7DCC2", "accent": "#34A853"},
    "material": {"surface": "#FAF6FE", "header": "#F4ECFC", "chip": "#F3E8FD", "border": "#DCC2F3", "accent": "#A142F4"},
}
SOURCE_TONES_DARK = {
    "line1": {"surface": "#142033", "header": "#182942", "chip": "#1D3557", "border": "#315A8F", "accent": "#8AB4F8"},
    "line2": {"surface": "#291A1C", "header": "#342024", "chip": "#46272B", "border": "#744046", "accent": "#F28B82"},
    "seonjin": {"surface": "#282317", "header": "#342D1B", "chip": "#463B20", "border": "#756334", "accent": "#FDD663"},
    "superlarge": {"surface": "#17261D", "header": "#1D3024", "chip": "#24402D", "border": "#3D6B49", "accent": "#81C995"},
    "material": {"surface": "#251B2F", "header": "#30223D", "chip": "#402B54", "border": "#674886", "accent": "#D7AEFB"},
}
COLORS = dict(LIGHT_COLORS)
FONTS = {
    "title": ("Malgun Gothic", 23, "bold"),
    "h2": ("Malgun Gothic", 13, "bold"),
    "body": ("Malgun Gothic", 10),
    "body_bold": ("Malgun Gothic", 10, "bold"),
    "small": ("Malgun Gothic", 9),
    "small_bold": ("Malgun Gothic", 9, "bold"),
    "metric": ("Malgun Gothic", 17, "bold"),
}
SPACING = {"xs": 4, "sm": 8, "md": 12, "lg": 16, "xl": 24}
def set_color_mode(is_dark: bool) -> None:
    COLORS.clear()
    COLORS.update(DARK_COLORS if is_dark else LIGHT_COLORS)


def source_tone(source_key: str, is_dark: bool = False) -> dict[str, str]:
    palette = SOURCE_TONES_DARK if is_dark else SOURCE_TONES_LIGHT
    return palette[source_key]

def status_meta(status: str, is_dark: bool = False):
    if status == "완료":
        return (COLORS["success"], "#064E3B" if is_dark else "#DCFCE7")
    if status == "오류":
        return (COLORS["danger"], "#7F1D1D" if is_dark else "#FEE2E2")
    return (COLORS["muted"], COLORS["surface_alt"])

def clean(value) -> str:
    return str(value or "").replace(" ", "").strip()


def default_columns_for_source(source_key: str) -> tuple[str, ...]:
    return MATERIAL_DEFAULT_COLUMNS if source_key == "material" else MONTHLY_DEFAULT_COLUMNS


def default_headers_for_source(source_key: str, headers: list[str]) -> list[str]:
    targets = {clean(column).upper() for column in default_columns_for_source(source_key)}
    return [header for header in headers if clean(header).upper() in targets]


def missing_default_columns(source_key: str, headers: list[str]) -> list[str]:
    available = {clean(header).upper() for header in headers}
    return [column for column in default_columns_for_source(source_key) if clean(column).upper() not in available]


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def display_value(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return text(value)


def visual_text_width(value) -> int:
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1 for char in str(value or ""))


def compact_filename(name: str, max_length: int = 30) -> str:
    if len(name) <= max_length:
        return name
    suffix = Path(name).suffix
    stem = Path(name).stem
    available = max(max_length - len(suffix) - 3, 8)
    head_length = max(int(available * 0.65), 4)
    tail_length = max(available - head_length, 4)
    return f"{stem[:head_length]}...{stem[-tail_length:]}{suffix}"


def sort_date_value(value) -> str:
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%Y-%m-%d")
        except Exception:
            pass
    value = display_value(value)
    if not value:
        return "9999-12-31"
    for sep in ("-", ".", "/"):
        parts = value.split(sep)
        if len(parts) >= 3 and len(parts[0]) == 4:
            return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2][:2].zfill(2)}"
    return value


def filename_start_date_range(values: list[str]) -> str:
    parsed = []
    for value in values:
        try:
            parsed.append(datetime.strptime(str(value), "%Y-%m-%d"))
        except (TypeError, ValueError):
            continue
    if not parsed:
        return "착수일미확인"
    return f"{min(parsed):%m%d}-{max(parsed):%m%d}"

def score_header(values, required) -> int:
    cells = {clean(value).upper() for value in values}
    return sum(1 for header in required if clean(header).upper() in cells)


def detect_header_row(ws, preferred: int, required: list[str]) -> int:
    best_row = preferred
    best_score = score_header([cell.value for cell in ws[preferred]], required)
    for row_no in range(1, min(ws.max_row, 15) + 1):
        score = score_header([cell.value for cell in ws[row_no]], required)
        if score > best_score:
            best_row, best_score = row_no, score
    return best_row


def unique_headers(headers: list[str]) -> list[str]:
    seen = Counter()
    result = []
    for index, header in enumerate(headers, 1):
        value = clean(header) or f"빈컬럼{index}"
        seen[value] += 1
        result.append(value if seen[value] == 1 else f"{value}_{seen[value]}")
    return result


def excel_com_convert_to_xlsx(path: Path) -> Path:
    if os.name != "nt":
        raise RuntimeError('Excel 자동 변환은 Windows에서만 사용할 수 있습니다.')
    output = Path(tempfile.gettempdir()) / f"wonwoo_excel_convert_{os.getpid()}_{datetime.now():%Y%m%d%H%M%S%f}.xlsx"
    script = f"""
$ErrorActionPreference = 'Stop'
$src = '{str(path).replace("'", "''")}'
$dst = '{str(output).replace("'", "''")}'
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($src, 0, $true)
    $workbook.SaveAs($dst, 51)
}} finally {{
    if ($workbook -ne $null) {{ $workbook.Close($false) | Out-Null }}
    if ($excel -ne $null) {{ $excel.Quit() | Out-Null }}
    [System.GC]::Collect()
    [System.GC]::WaitForPendingFinalizers()
}}
"""
    encoded = base64.b64encode(script.encode("utf-16le")).decode("ascii")
    command = ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-EncodedCommand", encoded]
    completed = subprocess.run(command, capture_output=True, text=True, timeout=120, creationflags=subprocess.CREATE_NO_WINDOW)
    if completed.returncode != 0 or not output.exists():
        detail = (completed.stderr or completed.stdout or "").strip()
        raise RuntimeError(detail or 'Excel 변환 결과 파일을 생성하지 못했습니다.')
    return output


def open_workbook_with_fallback(path: Path):
    try:
        return load_workbook(path, data_only=True), None
    except BadZipFile as exc:
        try:
            converted = excel_com_convert_to_xlsx(path)
            return load_workbook(converted, data_only=True), converted
        except Exception as fallback_exc:
            raise ValueError(
                '엑셀 파일을 직접 읽지 못했습니다. 파일이 DRM 보안 처리되었거나 표준 xlsx 구조가 아닐 수 있습니다. '
                'PC에 Microsoft Excel이 설치되어 있으면 자동 변환을 시도하지만 실패했습니다. '
                "Excel에서 파일을 열어 '다른 이름으로 저장(.xlsx)' 후 다시 선택하세요. "
                f"원본 오류: {exc}; 변환 오류: {fallback_exc}"
            ) from fallback_exc


def read_source(path: Path, preferred_header_row: int, required: list[str]) -> dict:
    if not path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
    if path.name.startswith("~$"):
        raise ValueError('Excel 임시 잠금 파일(~$)은 선택할 수 없습니다. 원본 엑셀 파일을 선택하세요.')
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError('.xlsx 또는 .xlsm 파일만 지원합니다. .xls 파일은 xlsx로 저장 후 선택하세요.')
    wb, converted_path = open_workbook_with_fallback(path)
    try:
        ws = wb.active
        header_row = detect_header_row(ws, preferred_header_row, required)
        raw_headers = [cell.value for cell in ws[header_row]]
        last_header_index = max((index for index, header in enumerate(raw_headers) if clean(header)), default=-1)
        if last_header_index < 0:
            raise ValueError('헤더가 비어 있습니다.')
        headers = unique_headers(raw_headers[: last_header_index + 1])
        rows = []
        for row_no, values in enumerate(ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True), header_row + 1):
            values = list(values)
            if any(text(value) for value in values):
                rows.append({"row_number": row_no, "values": values})
    finally:
        wb.close()
        if converted_path:
            converted_path.unlink(missing_ok=True)
    return {"path": path, "headers": headers, "rows": rows, "header_row": header_row}

def find_col(headers: list[str], names: list[str]) -> int | None:
    targets = {clean(name).upper() for name in names}
    for index, header in enumerate(headers):
        if clean(header).upper() in targets:
            return index
    return None


def value_at(row: dict, index: int | None):
    if index is None or index >= len(row["values"]):
        return ""
    return row["values"][index]


def key_value(row: dict, index: int | None) -> str:
    return clean(value_at(row, index))


def is_counterweight_name(value) -> bool:
    normalized = str(value or "").replace(" ", "").upper()
    return "카운터웨이트" in normalized or "COUNTERWEIGHT" in normalized


def exception(kind, source, key, reason) -> dict:
    return {"예외유형": kind, "파일구분": source, "Key": key, "사유": reason}


def model_exclusion(source_key: str, model_value) -> tuple[str, str] | None:
    model = clean(model_value).upper()
    if source_key == "superlarge" and model.startswith(("HX1000", "DX1000", "DX800")):
        return (
            "초대형 비대상 모델 제외",
            "초대형 라인 HX1000, DX1000, DX800 모델은 원우ENG 생산 카운터웨이트 대상이 아니므로 병합에서 제외합니다.",
        )
    if model.startswith("DX"):
        return (
            "DX 제품 제외",
            f"{SOURCE_LABELS[source_key]}의 DX~ 디벨론 제품은 원우ENG 자재소요현황 조회 대상에서 제외합니다.",
        )
    if source_key == "seonjin" and ("ACR" in model or "ECR" in model):
        return (
            "ACR/ECR 제품 제외",
            "선진정공 ACR(주물 카운터웨이트), ECR(전기굴착기) 모델은 원우ENG 생산 카운터웨이트 대상이 아니므로 병합에서 제외합니다.",
        )
    return None


def choose_material_row(key: str, rows: list[dict], material_name_index: int | None, exceptions: list[dict]) -> dict | None:
    if len(rows) <= 1:
        return rows[0] if rows else None
    cwt_rows = [row for row in rows if is_counterweight_name(value_at(row, material_name_index))]
    if len(cwt_rows) == 1:
        exceptions.append(exception("Key 중복-CWT 선택", "자재소요현황", key, f"동일 물류번호가 {len(rows)}건 존재하여 품명에 카운터웨이트/COUNTERWEIGHT가 포함된 행을 사용했습니다."))
        return cwt_rows[0]
    if len(cwt_rows) > 1:
        exceptions.append(exception("Key 중복-CWT 중복", "자재소요현황", key, f"동일 물류번호에 카운터웨이트 자재가 {len(cwt_rows)}건 존재합니다. 첫 번째 카운터웨이트 행을 사용합니다."))
        return cwt_rows[0]
    names = " / ".join(sorted({display_value(value_at(row, material_name_index)) for row in rows if display_value(value_at(row, material_name_index))}))
    exceptions.append(exception("Key 중복-CWT 없음", "자재소요현황", key, f"동일 물류번호가 {len(rows)}건 존재하지만 품명에서 카운터웨이트/COUNTERWEIGHT를 찾지 못해 제외했습니다. 품명: {names}"))
    return None


def canonical_output_name(header: str) -> str | None:
    normalized = clean(header).upper()
    if normalized in {"생산번호", "물류번호"}:
        return "생산번호"
    if normalized in {"착수일", "착수일자"}:
        return "착수일자"
    if normalized == "RADAR":
        return "Radar"
    return None


def output_column_name(prefix: str, header: str, duplicate_names: set[str]) -> str:
    canonical = canonical_output_name(header)
    if canonical:
        return canonical
    return f"{prefix}_{header}" if header in duplicate_names else header


def build_selected_indices(headers: list[str], selected_headers: list[str], key_index: int | None) -> list[int]:
    selected = set(selected_headers)
    indices = []
    if key_index is not None:
        indices.append(key_index)
    for index, header in enumerate(headers):
        if header in selected and index not in indices:
            indices.append(index)
    return indices


def index_materials(material_source: dict, key_index: int | None, exceptions: list[dict]) -> dict[str, list[dict]]:
    indexed = defaultdict(list)
    for row in material_source["rows"]:
        key = key_value(row, key_index)
        if not key:
            exceptions.append(exception("필수 Key 누락", "자재소요현황", "", f"{row['row_number']}행 물류번호가 비어 있습니다."))
            continue
        indexed[key].append(row)
    return indexed


def merge_sources(monthly_sources: dict[str, dict | None], material_source: dict, selections: dict[str, list[str]], output_dir: Path) -> Path:
    exceptions = []
    monthly_available = [(key, monthly_sources.get(key)) for key in MONTHLY_SOURCE_KEYS if monthly_sources.get(key)]
    if not monthly_available:
        raise ValueError("월확정서열 통합1라인, 통합2라인, 선진정공, 초대형 파일 중 하나 이상을 선택하세요.")
    if not material_source:
        raise ValueError("물류번호별 자재소요현황 파일을 선택하세요.")

    material_key = find_col(material_source["headers"], ["물류번호", "생산번호"])
    material_start_key = find_col(material_source["headers"], ["착수일", "착수일자"])
    material_name_key = find_col(material_source["headers"], ["품명"])
    if material_key is None:
        raise ValueError("자재소요현황에서 물류번호 컬럼을 찾을 수 없습니다.")

    monthly_key_by_source = {}
    monthly_model_by_source = {}
    monthly_start_by_source = {}
    for key, source in monthly_available:
        key_index = find_col(source["headers"], ["생산번호", "물류번호"])
        if key_index is None:
            raise ValueError(f"{SOURCE_LABELS[key]}에서 생산번호 컬럼을 찾을 수 없습니다.")
        monthly_key_by_source[key] = key_index
        monthly_model_by_source[key] = find_col(source["headers"], ["영업모델"])
        monthly_start_by_source[key] = find_col(source["headers"], ["착수일", "착수일자"])

    selected_monthly_headers = []
    for key, _source in monthly_available:
        selected_monthly_headers.extend(selections.get(key, []))
    selected_material_headers = selections.get("material", [])
    duplicate_names = set(selected_monthly_headers) & set(selected_material_headers)

    monthly_indices = {
        key: build_selected_indices(source["headers"], selections.get(key, []), monthly_key_by_source[key])
        for key, source in monthly_available
    }
    material_indices = build_selected_indices(material_source["headers"], selected_material_headers, material_key)
    if material_start_key is not None and material_start_key not in material_indices:
        material_indices.append(material_start_key)

    output_headers = ["조립라인"]
    monthly_header_pairs = []
    for key, source in monthly_available:
        for index in monthly_indices[key]:
            header = source["headers"][index]
            out_name = output_column_name("월확정", header, duplicate_names)
            if out_name not in output_headers:
                output_headers.append(out_name)
            monthly_header_pairs.append((key, index, out_name))

    material_header_pairs = []
    for index in material_indices:
        header = material_source["headers"][index]
        out_name = output_column_name("자재", header, duplicate_names)
        if out_name not in output_headers:
            output_headers.append(out_name)
            material_header_pairs.append((index, out_name))
        elif canonical_output_name(header) is None:
            material_header_pairs.append((index, out_name))

    if "비고" not in output_headers:
        output_headers.append("비고")

    material_indexed = index_materials(material_source, material_key, exceptions)
    known_monthly_keys = set()
    included_monthly_rows = []
    monthly_key_occurrences = defaultdict(list)
    exclusion_counts = Counter()

    for source_key, source in monthly_available:
        for row in source["rows"]:
            key = key_value(row, monthly_key_by_source[source_key])
            if not key:
                exceptions.append(exception("필수 Key 누락", SOURCE_LABELS[source_key], "", f"{row['row_number']}행 생산번호가 비어 있습니다."))
                continue
            known_monthly_keys.add(key)
            model = value_at(row, monthly_model_by_source.get(source_key))
            excluded = model_exclusion(source_key, model)
            if excluded:
                kind, reason = excluded
                exceptions.append(exception(kind, SOURCE_LABELS[source_key], key, reason))
                exclusion_counts[(source_key, kind)] += 1
                continue
            included_monthly_rows.append((source_key, source, row, key))
            monthly_key_occurrences[key].append((source_key, row["row_number"]))

    for key, occurrences in monthly_key_occurrences.items():
        if len(occurrences) > 1:
            locations = ", ".join(f"{SOURCE_LABELS[source_key]} {row_no}행" for source_key, row_no in occurrences)
            exceptions.append(exception("월확정서열 Key 중복", "월확정조립서열계획", key, f"동일 생산번호가 {len(occurrences)}건 존재합니다: {locations}. 각 행을 모두 병합합니다."))

    merged_rows = []
    plan_start_dates = []
    for source_key, source, row, key in included_monthly_rows:
        monthly_start = sort_date_value(value_at(row, monthly_start_by_source.get(source_key)))
        if monthly_start != "9999-12-31":
            plan_start_dates.append(monthly_start)
        matches = material_indexed.get(key, [])
        material_row = None
        remark = ""
        if not matches:
            remark = "생산번호와 일치하는 물류번호가 자재소요현황에 없습니다"
        else:
            material_row = choose_material_row(key, matches, material_name_key, exceptions)
            if material_row is None:
                continue
        output = {header: "" for header in output_headers}
        output["조립라인"] = SOURCE_LABELS[source_key].replace("월확정서열 ", "")
        for pair_key, index, out_name in monthly_header_pairs:
            if pair_key == source_key:
                output[out_name] = display_value(value_at(row, index))
        if material_row is not None:
            for index, out_name in material_header_pairs:
                output[out_name] = display_value(value_at(material_row, index))
        if remark:
            output["비고"] = remark
        output["__sort_start_date"] = monthly_start
        output["__sort_source"] = SOURCE_LABELS[source_key]
        output["__sort_key"] = key
        merged_rows.append(output)

    for key in material_indexed:
        if key not in known_monthly_keys:
            exceptions.append(exception("월확정서열 미매칭", "자재소요현황", key, "물류번호와 일치하는 월확정서열 생산번호가 없습니다."))

    merged_rows.sort(key=lambda row: (row.get("__sort_start_date", "9999-12-31"), row.get("__sort_source", ""), row.get("__sort_key", "")))
    for row in merged_rows:
        row.pop("__sort_start_date", None)
        row.pop("__sort_source", None)
        row.pop("__sort_key", None)

    creation_date = datetime.now().strftime("%Y%m%d")
    start_date_range = filename_start_date_range(plan_start_dates)
    output_path = output_dir / f"원우ENG_Key_Column_Merge_{creation_date}_{start_date_range}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "병합결과", merged_rows or [{header: "" for header in output_headers}], output_headers)
    write_sheet(wb, "예외목록", exceptions or [exception("예외 없음", "", "", "예외가 없습니다.")], EXCEPTION_COLUMNS)
    summary = [["구분", "건수", "비고"]]
    available_map = dict(monthly_available)
    for key in MONTHLY_SOURCE_KEYS:
        source = available_map.get(key)
        if source:
            summary.append([SOURCE_LABELS[key], len(source["rows"]), source["path"].name])
            summary.append([f"{SOURCE_LABELS[key]} 선택 컬럼", len(selections.get(key, [])), ", ".join(selections.get(key, []))])
            summary.append([f"{SOURCE_LABELS[key]} 모델 제외", sum(count for (source_key, _kind), count in exclusion_counts.items() if source_key == key), "DX~/WX~ 제외 규칙"])
        else:
            summary.append([SOURCE_LABELS[key], 0, "미선택"])
    summary.append(["자재소요현황", len(material_source["rows"]), material_source["path"].name])
    summary.append(["자재소요현황 선택 컬럼", len(selected_material_headers), ", ".join(selected_material_headers)])
    summary.append(["DX 제품 제외", sum(count for (_key, kind), count in exclusion_counts.items() if kind == "DX 제품 제외"), "전체 월확정서열 라인"])
    summary.append(["정상 병합", len(merged_rows), "병합결과 시트"])
    summary.append(["예외", len(exceptions), "예외목록 시트"])
    summary.append(["앱 버전", APP_VERSION, "Python + openpyxl + tkinter"])
    write_sheet(wb, "검증요약", summary)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path

def write_sheet(wb, title, rows, columns=None):
    ws = wb.create_sheet(title)
    if rows and isinstance(rows[0], dict):
        headers = columns or list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
    else:
        for row in rows or []:
            ws.append(row)
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    for cells in ws.columns:
        width = min(max(visual_text_width(cell.value) for cell in cells) + 2, 48)
        ws.column_dimensions[cells[0].column_letter].width = max(width, 10)


def open_folder(path: Path) -> None:
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
    except Exception:
        pass


def set_console_title(title: str) -> None:
    if not sys.platform.startswith("win"):
        return
    try:
        import ctypes
        ctypes.windll.kernel32.SetConsoleTitleW(title)
    except Exception:
        pass


class ToolTip:
    def __init__(self, widget, text_value: str = ""):
        self.widget = widget
        self.text_value = text_value
        self.window = None
        widget.bind("<Enter>", self._show, add="+")
        widget.bind("<Leave>", self._hide, add="+")

    def set_text(self, text_value: str):
        self.text_value = text_value

    def _show(self, _event=None):
        if not self.text_value or self.window:
            return
        x = self.widget.winfo_rootx() + 8
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self.window = tk.Toplevel(self.widget)
        self.window.wm_overrideredirect(True)
        self.window.wm_geometry(f"+{x}+{y}")
        tk.Label(
            self.window,
            text=self.text_value,
            bg="#111827",
            fg="#FFFFFF",
            font=FONTS["small"],
            padx=9,
            pady=6,
            relief="solid",
            bd=1,
        ).pack()

    def _hide(self, _event=None):
        if self.window:
            self.window.destroy()
            self.window = None

class ScrollableCheckList(tk.Frame):
    def __init__(self, parent, title: str, tone: dict[str, str], on_change=None):
        super().__init__(parent, bg=tone["surface"], highlightbackground=tone["border"], highlightthickness=1)
        self.title = title
        self.tone = tone
        self.on_change = on_change
        self.vars: dict[str, tk.BooleanVar] = {}
        self.all_headers: list[str] = []
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_args: self._render_headers())

        header = tk.Frame(self, bg=tone["header"], padx=SPACING["md"], pady=SPACING["sm"])
        header.pack(fill="x")
        tk.Label(header, text=title, width=1, anchor="w", bg=tone["header"], fg=tone["accent"], font=FONTS["body_bold"]).pack(side="left", fill="x", expand=True)
        self.count_label = tk.Label(header, text="선택 0개", bg=tone["chip"], fg=tone["accent"], font=FONTS["small_bold"], padx=8, pady=3)
        self.count_label.pack(side="right")

        controls = tk.Frame(self, bg=tone["surface"], padx=SPACING["sm"])
        controls.pack(fill="x", pady=(SPACING["sm"], SPACING["sm"]))
        search = tk.Entry(controls, textvariable=self.search_var, bg=tone["header"], fg=COLORS["text"], insertbackground=COLORS["text"], relief="flat", font=FONTS["small"])
        search.insert(0, "")
        search.pack(fill="x", ipady=5)
        button_row = tk.Frame(controls, bg=tone["surface"])
        button_row.pack(fill="x", pady=(SPACING["xs"], 0))
        self._small_button(button_row, "전체 선택", self.select_all).pack(side="left", fill="x", expand=True, padx=(0, 2))
        self._small_button(button_row, "전체 해제", self.clear_all).pack(side="left", fill="x", expand=True, padx=(2, 0))

        self.canvas = tk.Canvas(self, width=150, height=258, bg=tone["surface"], highlightthickness=0)
        self.inner = tk.Frame(self.canvas, bg=tone["surface"])
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True, padx=(SPACING["md"], 0), pady=(0, SPACING["md"]))
        self.scrollbar.pack(side="right", fill="y", padx=(0, SPACING["md"]), pady=(0, SPACING["md"]))
        self.window_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.inner.bind("<Configure>", lambda _e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfigure(self.window_id, width=e.width))
        for widget in (self.canvas, self.inner):
            widget.bind("<MouseWheel>", self._on_mousewheel)
            widget.bind("<Button-4>", self._on_mousewheel)
            widget.bind("<Button-5>", self._on_mousewheel)

    def _small_button(self, parent, text_value: str, command):
        return tk.Button(parent, text=text_value, command=command, bg=self.tone["chip"], fg=self.tone["accent"], activebackground=self.tone["header"], activeforeground=self.tone["accent"], relief="solid", bd=1, highlightbackground=self.tone["border"], highlightthickness=0, font=FONTS["small_bold"], padx=8, pady=5, cursor="hand2")

    def _on_mousewheel(self, event):
        if getattr(event, "num", None) == 4:
            delta = -3
        elif getattr(event, "num", None) == 5:
            delta = 3
        else:
            delta = -1 * int(event.delta / 120) if event.delta else 0
        if delta:
            self.canvas.yview_scroll(delta, "units")
        return "break"

    def _render_headers(self):
        for child in self.inner.winfo_children():
            child.destroy()
        keyword = clean(self.search_var.get()).upper()
        visible = [header for header in self.all_headers if keyword in clean(header).upper()]
        if not visible:
            tk.Label(self.inner, text="표시할 컬럼이 없습니다.", bg=self.tone["surface"], fg=COLORS["muted"], font=FONTS["small"], pady=12).pack(anchor="w", padx=8)
        for header in visible:
            var = self.vars[header]
            row = tk.Frame(self.inner, bg=self.tone["surface"], pady=1)
            row.pack(fill="x", padx=4, pady=1)
            cb = tk.Checkbutton(row, text=header, width=1, variable=var, bg=self.tone["surface"], fg=COLORS["text"], activebackground=self.tone["header"], activeforeground=COLORS["text"], selectcolor=self.tone["surface"], anchor="w", justify="left", font=FONTS["small"], command=self._selection_changed, wraplength=155)
            cb.pack(fill="x", anchor="w")
            for widget in (row, cb):
                widget.bind("<MouseWheel>", self._on_mousewheel)
                widget.bind("<Button-4>", self._on_mousewheel)
                widget.bind("<Button-5>", self._on_mousewheel)
        self._selection_changed()

    def set_headers(self, headers: list[str], default_selected: list[str] | None = None):
        self.all_headers = headers
        self.vars.clear()
        defaults = set(default_selected or [])
        for header in headers:
            var = tk.BooleanVar(value=header in defaults)
            var.trace_add("write", lambda *_args: self._selection_changed())
            self.vars[header] = var
        self.search_var.set("")
        self._render_headers()

    def selected(self) -> list[str]:
        return [header for header, var in self.vars.items() if var.get()]

    def _selection_changed(self):
        selected_count = len(self.selected())
        total_count = len(self.vars)
        self.count_label.configure(text=f"선택 {selected_count}/{total_count}")
        if self.on_change:
            self.on_change()

    def select_all(self):
        for var in self.vars.values():
            var.set(True)
        self._selection_changed()

    def clear_all(self):
        for var in self.vars.values():
            var.set(False)
        self._selection_changed()

class MergeGui(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1360x900")
        self.minsize(980, 560)
        self.configure(bg=COLORS["bg"])
        self.dark_mode = tk.BooleanVar(value=False)
        self.files: dict[str, Path | None] = {**{key: None for key in ALL_SOURCE_KEYS}, "output_dir": Path.home() / "Desktop"}
        self.sources: dict[str, dict | None] = {key: None for key in ALL_SOURCE_KEYS}
        self.last_output: Path | None = None
        self.file_labels: dict[str, tk.Label] = {}
        self.file_panels: dict[str, tk.Frame] = {}
        self.file_tooltips: dict[str, ToolTip] = {}
        self.file_status: dict[str, tk.Label] = {}
        self.file_meta: dict[str, tk.Label] = {}
        self.metric_labels: dict[str, tk.Label] = {}
        self.checklists: dict[str, ScrollableCheckList] = {}
        self._build_styles()
        self._build_ui()
        self._update_state()

    def _build_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(
            "TScrollbar",
            background=COLORS["panel"],
            troughcolor=COLORS["surface_alt"],
            bordercolor=COLORS["border"],
            arrowcolor=COLORS["muted"],
        )

        if self.dark_mode.get():
            page_thumb = "#60A5FA"
            page_active = "#93C5FD"
            page_pressed = "#3B82F6"
            page_trough = "#263449"
            page_border = "#64748B"
            page_arrow = "#0B1120"
        else:
            page_thumb = "#2563EB"
            page_active = "#1D4ED8"
            page_pressed = "#1E40AF"
            page_trough = "#CBD5E1"
            page_border = "#94A3B8"
            page_arrow = "#FFFFFF"

        style.configure(
            "Page.Vertical.TScrollbar",
            background=page_thumb,
            troughcolor=page_trough,
            bordercolor=page_border,
            lightcolor=page_thumb,
            darkcolor=page_thumb,
            arrowcolor=page_arrow,
            borderwidth=1,
            arrowsize=18,
        )
        style.map(
            "Page.Vertical.TScrollbar",
            background=[("pressed", page_pressed), ("active", page_active)],
            arrowcolor=[("pressed", page_arrow), ("active", page_arrow)],
        )

    def _build_ui(self):
        shell = tk.Frame(self, bg=COLORS["bg"])
        shell.pack(fill="both", expand=True)
        self.page_canvas = tk.Canvas(shell, bg=COLORS["bg"], highlightthickness=0)
        self.page_scrollbar = ttk.Scrollbar(shell, orient="vertical", command=self.page_canvas.yview, style="Page.Vertical.TScrollbar")
        self.page_canvas.configure(yscrollcommand=self.page_scrollbar.set)
        self.page_canvas.pack(side="left", fill="both", expand=True)
        self.page_scrollbar.pack(side="right", fill="y")

        root = tk.Frame(self.page_canvas, bg=COLORS["bg"], padx=22, pady=18)
        self.page_window_id = self.page_canvas.create_window((0, 0), window=root, anchor="nw")
        root.bind("<Configure>", self._on_page_configure)
        self.page_canvas.bind("<Configure>", self._on_canvas_configure)
        for widget in (self, self.page_canvas, root):
            widget.bind("<MouseWheel>", self._on_page_mousewheel)
            widget.bind("<Button-4>", self._on_page_mousewheel)
            widget.bind("<Button-5>", self._on_page_mousewheel)

        self._build_header(root)
        self._build_file_flow(root)
        self._build_metrics(root)
        self._build_column_area(root)
        self._build_action_area(root)

    def _on_page_configure(self, _event=None):
        self.page_canvas.configure(scrollregion=self.page_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.page_canvas.itemconfigure(self.page_window_id, width=event.width)

    def _on_page_mousewheel(self, event):
        if getattr(event, "num", None) == 4:
            delta = -3
        elif getattr(event, "num", None) == 5:
            delta = 3
        else:
            delta = -1 * int(event.delta / 120) if event.delta else 0
        if delta:
            self.page_canvas.yview_scroll(delta, "units")
        return "break"

    def _build_header(self, parent):
        header = tk.Frame(parent, bg=COLORS["dark"], padx=26, pady=20)
        header.pack(fill="x", pady=(0, SPACING["lg"]))
        left = tk.Frame(header, bg=COLORS["dark"])
        left.pack(side="left", fill="both", expand=True)
        tk.Label(left, text=APP_TITLE, bg=COLORS["dark"], fg=COLORS["on_dark"], font=FONTS["title"]).pack(anchor="w")
        tk.Label(left, text="월확정서열 서열정보와 소요자재를 생산번호 기준으로 자동 취합하고, 선택 컬럼만 엑셀로 생성합니다.", bg=COLORS["dark"], fg=COLORS["header_sub"], font=FONTS["body"]).pack(anchor="w", pady=(8, 0))
        toggle_row = tk.Frame(left, bg=COLORS["dark"])
        toggle_row.pack(anchor="w", pady=(14, 0))
        tk.Checkbutton(toggle_row, text="다크모드", variable=self.dark_mode, command=self.toggle_theme, bg=COLORS["dark"], fg=COLORS["on_dark"], activebackground=COLORS["dark_panel"], activeforeground=COLORS["on_dark"], selectcolor=COLORS["dark_panel"], font=FONTS["small_bold"], cursor="hand2").pack(side="left")
        tk.Label(toggle_row, text="화면 대비를 높여 장시간 작업 시 눈부심을 줄입니다.", bg=COLORS["dark"], fg=COLORS["header_sub"], font=FONTS["small"]).pack(side="left", padx=(10, 0))
        right = tk.Frame(header, bg=COLORS["dark_panel"], padx=16, pady=12, highlightbackground="#334155", highlightthickness=1)
        right.pack(side="right", padx=(18, 0))
        tk.Label(right, text="기준 KEY", bg=COLORS["dark_panel"], fg=COLORS["header_key"], font=FONTS["small_bold"]).pack(anchor="w")
        tk.Label(right, text="생산번호 = 물류번호", bg=COLORS["dark_panel"], fg=COLORS["on_dark"], font=("Malgun Gothic", 13, "bold")).pack(anchor="w", pady=(4, 10))
        tk.Label(right, text="제작: HD건설기계 협력사육성팀 조병철 선임매니저", bg=COLORS["dark_panel"], fg=COLORS["header_sub"], font=FONTS["small"]).pack(anchor="w")

    def _build_file_flow(self, parent):
        section = tk.Frame(parent, bg=COLORS["bg"])
        section.pack(fill="x", pady=(0, SPACING["md"]))
        tk.Label(section, text="1. 월확정서열 조립라인 및 자재 파일 선택", bg=COLORS["bg"], fg=COLORS["text"], font=FONTS["h2"]).pack(anchor="w", pady=(0, SPACING["sm"]))

        cards = tk.Frame(section, bg=COLORS["bg"])
        cards.pack(fill="x")
        for column in range(5):
            cards.columnconfigure(column, weight=1, uniform="source_file_cards")

        sources = (
            ("line1", "1", "월확정서열 통합1라인", "선택 사항 · 생산번호 기준"),
            ("line2", "2", "월확정서열 통합2라인", "선택 사항 · DX~ 제외"),
            ("seonjin", "3", "월확정서열 선진정공", "선택 사항 · DX~/WX~ 제외"),
            ("superlarge", "4", "월확정서열 초대형", "선택 사항 · DX~ 제외"),
            ("material", "5", "물류번호별 자재소요현황", "필수 · 물류번호 기준"),
        )
        for column, (key, number, title, hint) in enumerate(sources):
            left = 0 if column == 0 else SPACING["xs"]
            right = 0 if column == len(sources) - 1 else SPACING["xs"]
            self._file_panel(cards, key, number, title, hint, grid_column=column, padx=(left, right))

    def _build_metrics(self, parent):
        metrics = tk.Frame(parent, bg=COLORS["bg"])
        metrics.pack(fill="x", pady=(0, SPACING["md"]))
        metric_items = (("uploads", "업로드 현황"), ("columns", "선택 컬럼"), ("normal", "정상병합"), ("exceptions", "예외 수"), ("result", "생성 결과"))
        for column in range(len(metric_items)):
            metrics.columnconfigure(column, weight=1, uniform="summary_metric_cards")

        for column, (key, title) in enumerate(metric_items):
            card = tk.Frame(metrics, bg=COLORS["surface"], padx=16, pady=12, highlightbackground=COLORS["border"], highlightthickness=1)
            left = 0 if column == 0 else SPACING["xs"]
            right = 0 if column == len(metric_items) - 1 else SPACING["xs"]
            card.grid(row=0, column=column, sticky="nsew", padx=(left, right))
            tk.Label(card, text=title, width=1, anchor="w", bg=COLORS["surface"], fg=COLORS["muted"], font=FONTS["small_bold"]).pack(fill="x")
            value = tk.Label(card, text="-", width=1, anchor="w", justify="left", bg=COLORS["surface"], fg=COLORS["primary"], font=FONTS["metric"])
            value.pack(fill="x", pady=(4, 0))
            if key == "uploads":
                card.bind("<Configure>", lambda event, label=value: label.configure(wraplength=max(100, event.width - 34)), add="+")
            self.metric_labels[key] = value

    def _build_column_area(self, parent):
        box = tk.Frame(parent, bg=COLORS["surface"], padx=10, pady=12, highlightbackground=COLORS["border"], highlightthickness=1)
        box.pack(fill="both", expand=True, pady=(0, SPACING["md"]))

        title_row = tk.Frame(box, bg=COLORS["surface"])
        title_row.pack(fill="x", pady=(0, SPACING["sm"]))
        title_row.columnconfigure(1, weight=1)
        tk.Label(title_row, text="2. 출력 컬럼 선택", bg=COLORS["surface"], fg=COLORS["text"], font=FONTS["h2"]).grid(row=0, column=0, sticky="nw", padx=(0, SPACING["lg"]))

        monthly_defaults = " · ".join(MONTHLY_DEFAULT_COLUMNS)
        material_defaults = " · ".join(MATERIAL_DEFAULT_COLUMNS)
        default_note = (
            f"기본 선택 · 월확정서열(4개): {monthly_defaults}\n"
            f"자재소요현황: {material_defaults}"
        )
        note_label = tk.Label(
            title_row,
            text=default_note,
            width=1,
            bg=COLORS["surface"],
            fg=COLORS["muted"],
            font=FONTS["small"],
            anchor="e",
            justify="right",
        )
        note_label.grid(row=0, column=1, sticky="ew")
        title_row.bind(
            "<Configure>",
            lambda event: note_label.configure(wraplength=max(360, event.width - 230)),
            add="+",
        )

        lists = tk.Frame(box, bg=COLORS["surface"])
        lists.pack(fill="both", expand=True)
        lists.rowconfigure(0, weight=1)
        for index, key in enumerate(ALL_SOURCE_KEYS):
            lists.columnconfigure(index, weight=1, uniform="source_column_lists")
            checklist = ScrollableCheckList(lists, SOURCE_LABELS[key], source_tone(key, self.dark_mode.get()), on_change=self._update_state)
            checklist.grid(row=0, column=index, sticky="nsew", padx=SPACING["xs"])
            self.checklists[key] = checklist

    def _build_action_area(self, parent):
        action = tk.Frame(parent, bg=COLORS["surface"], padx=14, pady=12, highlightbackground=COLORS["border"], highlightthickness=1)
        action.pack(fill="x")
        self._result_step(action, "1", "폴더 지정", self.select_output_dir)
        self._flow_arrow(action, compact=True)
        self._result_step(action, "2", "병합 엑셀 생성", self.generate, primary=True)
        self._flow_arrow(action, compact=True)
        self._result_step(action, "3", "폴더 열기", self.open_result_folder)
        self._button(action, "초기화", self.reset, variant="secondary").pack(side="left", padx=(SPACING["md"], 0))
        right = tk.Frame(action, bg=COLORS["surface"])
        right.pack(side="right", fill="x", expand=True, padx=(SPACING["lg"], 0))
        self.status_var = tk.StringVar(value="월확정서열 파일 1개 이상과 자재소요현황 파일을 선택하세요.")
        tk.Label(right, textvariable=self.status_var, bg=COLORS["surface"], fg=COLORS["text"], font=FONTS["body_bold"], anchor="e").pack(fill="x")
        self.output_label = tk.Label(right, text=str(self.files["output_dir"]), bg=COLORS["surface"], fg=COLORS["muted"], font=FONTS["small"], anchor="e")
        self.output_label.pack(fill="x", pady=(4, 0))

    def _button(self, parent, text_value: str, command, variant: str = "primary"):
        if variant == "primary":
            bg, fg, active = COLORS["primary"], COLORS["on_dark"], COLORS["primary_hover"]
        else:
            bg, fg, active = COLORS["surface_alt"], COLORS["text"], COLORS["panel"]
        return tk.Button(parent, text=text_value, command=command, bg=bg, fg=fg, activebackground=active, activeforeground=fg, relief="flat", bd=0, highlightthickness=0, font=FONTS["body_bold"], padx=14, pady=9, cursor="hand2")

    def _status_badge(self, parent, status: str):
        fg, bg = status_meta(status, self.dark_mode.get())
        return tk.Label(parent, text=status, bg=bg, fg=fg, font=FONTS["body_bold"], padx=14, pady=9)

    def _file_panel(self, parent, key: str, number: str, title: str, hint: str, grid_column: int | None = None, padx=(0, 0)):
        tone = source_tone(key, self.dark_mode.get())
        panel = tk.Frame(parent, bg=tone["surface"], padx=14, pady=13, highlightbackground=tone["border"], highlightthickness=1)
        if grid_column is None:
            panel.pack(side="left", fill="x", expand=True, padx=padx)
        else:
            panel.grid(row=0, column=grid_column, sticky="nsew", padx=padx)
        self.file_panels[key] = panel

        top = tk.Frame(panel, bg=tone["header"])
        top.pack(fill="x")
        tk.Label(top, text=number, width=2, bg=tone["accent"], fg=COLORS["on_dark"], font=FONTS["body_bold"], padx=5, pady=4).pack(side="left")
        tk.Label(top, text=title, width=1, anchor="w", bg=tone["header"], fg=tone["accent"], font=FONTS["body_bold"]).pack(side="left", fill="x", expand=True, padx=(SPACING["sm"], SPACING["xs"]))

        tk.Label(panel, text=hint, bg=tone["surface"], fg=COLORS["muted"], font=FONTS["small"], anchor="w").pack(fill="x", pady=(8, 4))
        label = tk.Label(panel, text="파일을 선택하세요", width=1, bg=tone["surface"], fg=tone["accent"], font=FONTS["small_bold"], anchor="w", justify="left")
        label.pack(fill="x")
        self.file_labels[key] = label
        self.file_tooltips[key] = ToolTip(label)

        meta = tk.Label(panel, text="헤더/데이터 건수 대기", bg=tone["surface"], fg=COLORS["muted"], font=FONTS["small"], anchor="w")
        meta.pack(fill="x", pady=(4, 8))
        self.file_meta[key] = meta

        actions = tk.Frame(panel, bg=tone["surface"])
        actions.pack(fill="x")
        badge = self._status_badge(actions, "대기")
        badge.pack(side="left", fill="y")
        self.file_status[key] = badge
        select_button = self._button(actions, "파일 선택", lambda k=key: self.select_file(k), variant="secondary")
        select_button.configure(bg=tone["chip"], fg=tone["accent"], activebackground=tone["header"], activeforeground=tone["accent"])
        select_button.pack(side="right")
    def _flow_arrow(self, parent, compact: bool = False):
        try:
            bg = parent.cget("background")
        except tk.TclError:
            bg = COLORS["bg"]
        tk.Label(parent, text="→", bg=bg, fg=COLORS["secondary"], font=("Malgun Gothic", 20 if compact else 24, "bold")).pack(side="left", padx=(8, 8))

    def _result_step(self, parent, number: str, title: str, command, primary: bool = False):
        frame = tk.Frame(parent, bg=COLORS["surface_alt"], padx=8, pady=7, highlightbackground=COLORS["border_strong"], highlightthickness=1)
        frame.pack(side="left")
        tk.Label(frame, text=number, width=2, bg=COLORS["success"], fg=COLORS["on_dark"], font=FONTS["body_bold"], padx=4, pady=4).pack(side="left")
        button = self._button(frame, title, command, variant="primary" if primary else "secondary")
        button.pack(side="left", padx=(8, 0))
        if primary:
            self.merge_btn = button

    def _set_file_status(self, key: str, status: str):
        fg, bg = status_meta(status, self.dark_mode.get())
        self.file_status[key].configure(text=status, fg=fg, bg=bg)

    def _set_file_name(self, key: str, path: Path | None):
        if path is None:
            self.file_labels[key].configure(text="파일을 선택하세요")
            self.file_tooltips[key].set_text("")
            return
        self.file_labels[key].configure(text=compact_filename(path.name))
        self.file_tooltips[key].set_text(path.name)

    def toggle_theme(self):
        preserved = {key: checklist.selected() for key, checklist in self.checklists.items()} if self.checklists else {}
        last_output = self.last_output
        set_color_mode(self.dark_mode.get())
        for child in self.winfo_children():
            child.destroy()
        self.configure(bg=COLORS["bg"])
        self.file_labels.clear()
        self.file_panels.clear()
        self.file_tooltips.clear()
        self.file_status.clear()
        self.file_meta.clear()
        self.metric_labels.clear()
        self.checklists.clear()
        self.last_output = last_output
        self._build_styles()
        self._build_ui()
        for key in ALL_SOURCE_KEYS:
            source = self.sources.get(key)
            if source:
                self._set_file_status(key, "완료")
                self._set_file_name(key, source["path"])
                self.file_meta[key].configure(text=f"헤더 {source['header_row']}행 · 데이터 {len(source['rows'])}건")
                defaults = preserved[key] if key in preserved else default_headers_for_source(key, source["headers"])
                self.checklists[key].set_headers(source["headers"], defaults)
            else:
                self._set_file_status(key, "대기")
                self.checklists[key].set_headers([])
        self.output_label.configure(text=str(self.files["output_dir"]))
        self._update_state()
        if self.last_output:
            self.metric_labels["result"].configure(text=self.last_output.name)
            self.metric_labels["normal"].configure(text=f"{self._count_normal_merges(self.last_output)}건")
            self.metric_labels["exceptions"].configure(text=f"{self._count_exceptions(self.last_output)}건")
    def select_file(self, key: str):
        selected = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel Workbook", "*.xlsx *.xlsm"), ("All files", "*.*")])
        if not selected:
            return
        try:
            selected_path = Path(selected).resolve()
            duplicate_key = next(
                (
                    other_key
                    for other_key in ALL_SOURCE_KEYS
                    if other_key != key and self.files.get(other_key) and Path(self.files[other_key]).resolve() == selected_path
                ),
                None,
            )
            if duplicate_key:
                raise ValueError(f"같은 파일이 이미 {SOURCE_LABELS[duplicate_key]}에 선택되어 있습니다.")
            preferred = 1 if key == "material" else 2
            required = ["물류번호"] if key == "material" else ["생산번호"]
            source = read_source(selected_path, preferred, required)
        except Exception as exc:
            self._set_file_status(key, "오류")
            self.status_var.set(f"오류: {exc}")
            messagebox.showerror(APP_TITLE, str(exc))
            return
        self.files[key] = Path(selected)
        self.sources[key] = source
        self._set_file_status(key, "완료")
        self._set_file_name(key, Path(selected))
        self.file_meta[key].configure(text=f"헤더 {source['header_row']}행 · 데이터 {len(source['rows'])}건")
        defaults = default_headers_for_source(key, source["headers"])
        self.checklists[key].set_headers(source["headers"], defaults)
        self._update_state()

    def select_output_dir(self):
        selected = filedialog.askdirectory(title="결과 저장 폴더 선택")
        if selected:
            self.files["output_dir"] = Path(selected)
            self.output_label.configure(text=selected)
            self._update_state()

    def _update_state(self):
        monthly_done = sum(1 for key in MONTHLY_SOURCE_KEYS if self.sources[key])
        material_done = bool(self.sources["material"])
        ready = bool(monthly_done and material_done and self.files["output_dir"])
        if hasattr(self, "merge_btn"):
            self.merge_btn.configure(state="normal" if ready else "disabled", bg=COLORS["primary"] if ready else COLORS["panel"], fg=COLORS["on_dark"] if ready else COLORS["muted"], cursor="hand2" if ready else "arrow")
        selected_cols = sum(len(checklist.selected()) for checklist in self.checklists.values()) if self.checklists else 0
        if "uploads" in self.metric_labels:
            material_status = "완료" if material_done else "대기"
            self.metric_labels["uploads"].configure(text=f"월확정서열 {monthly_done}/4 · 자재 {material_status}")
        if "columns" in self.metric_labels:
            self.metric_labels["columns"].configure(text=f"{selected_cols}개")
        if "normal" in self.metric_labels:
            self.metric_labels["normal"].configure(text=f"{self._count_normal_merges(self.last_output)}건" if self.last_output else "생성 후 확인")
        if "exceptions" in self.metric_labels:
            self.metric_labels["exceptions"].configure(text=f"{self._count_exceptions(self.last_output)}건" if self.last_output else "생성 후 확인")
        if "result" in self.metric_labels:
            self.metric_labels["result"].configure(text=self.last_output.name if self.last_output else "대기")
        self.status_var.set("컬럼을 선택한 뒤 병합 엑셀 생성을 누르세요." if ready else "월확정서열 파일 1개 이상과 자재소요현황 파일을 선택하세요.")

    def _count_sheet_rows(self, output: Path, sheet_name: str, empty_marker: str | None = None) -> int:
        try:
            wb = load_workbook(output, read_only=True, data_only=True)
            ws = wb[sheet_name]
            count = max(ws.max_row - 1, 0)
            if count <= 0:
                wb.close()
                return 0
            first_values = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
            wb.close()
            if empty_marker and first_values and first_values[0] == empty_marker:
                return 0
            if all(value in (None, "") for value in first_values):
                return 0
            return count
        except Exception:
            return 0

    def _count_normal_merges(self, output: Path) -> int:
        return self._count_sheet_rows(output, "병합결과")

    def _count_exceptions(self, output: Path) -> int:
        return self._count_sheet_rows(output, "예외목록", "예외 없음")

    def generate(self):
        try:
            selections = {key: checklist.selected() for key, checklist in self.checklists.items()}
            self.status_var.set("병합 엑셀을 생성하는 중입니다...")
            self.update_idletasks()
            monthly_sources = {key: self.sources[key] for key in MONTHLY_SOURCE_KEYS}
            output = merge_sources(monthly_sources, self.sources["material"], selections, self.files["output_dir"])
        except Exception as exc:
            self.status_var.set(f"오류: {exc}")
            messagebox.showerror(APP_TITLE, str(exc))
            return
        self.last_output = output
        normal_count = self._count_normal_merges(output)
        exception_count = self._count_exceptions(output)
        self.metric_labels["normal"].configure(text=f"{normal_count}건")
        self.metric_labels["exceptions"].configure(text=f"{exception_count}건")
        self.metric_labels["result"].configure(text=output.name)
        self.status_var.set(f"완료: {output.name}")
        messagebox.showinfo(APP_TITLE, f"병합 엑셀 생성 완료\n\n{output}")

    def open_result_folder(self):
        folder = self.last_output.parent if self.last_output else self.files["output_dir"]
        if folder:
            open_folder(folder)

    def reset(self):
        self.files.update({key: None for key in ALL_SOURCE_KEYS})
        self.sources.update({key: None for key in ALL_SOURCE_KEYS})
        self.last_output = None
        for key in ALL_SOURCE_KEYS:
            self._set_file_status(key, "대기")
            self._set_file_name(key, None)
            self.file_meta[key].configure(text="헤더/데이터 건수 대기")
            self.checklists[key].set_headers([])
        self._update_state()

def main():
    set_console_title(APP_TITLE)
    app = MergeGui()
    app.mainloop()


if __name__ == "__main__":
    main()





























